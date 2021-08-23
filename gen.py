from openpyxl import Workbook
from openpyxl import load_workbook
import itertools
wb = load_workbook('stages.xlsx',)
print('sheets:',wb.sheetnames)


ws = wb['exportConvention']
names=ws['C']
firstnames = ws['D']
codes = ws['L']
sujets = ws['T']
entreprises = ws['BB']
www_entreprises = ws['BQ']

masterlist = (('m1', 'MI6251'), ('m2', 'MI6252'))

def writeRapports(f, n, fn, s, e, w ):
    f.write(
"""
 - [[[{0}]]] {1} {2}, _{4}_, link:{5}[{3}], link:{{attachmentsdir}}/++{1}-{2}.pdf++[{1}-{2}.pdf],  link:{{attachmentsdir}}/++{1}-{2}-slides.pdf++[{1}-{2}-slides.pdf] 
""".format(n.value.title().replace(" ", ""), n.value.title(), fn.value.title(), e.value.title().strip(), s.value.capitalize().strip(), w.value.strip()))


def writeTableEntry(f, n, fn, s, e, w):
    f.write(
        """
| {0} | {1} | link:{4}[{2}] | _{3}_
""".format(n.value.title(), fn.value.title(), e.value.title().strip(), s.value.strip(), w.value.strip()))


for module,master in masterlist:
    f = open("modules/"+module+"/partials/rapports.adoc", "w")
    for n, fn, c, s, e,w in sorted(zip(names, firstnames, codes, sujets, entreprises,www_entreprises), key=lambda x: x[0].value):
        if c.value == master :
            writeRapports(f, n, fn, s, e, w)
    f.close()
    f = open("modules/"+module+"/partials/stages.adoc", "w")
    f.write('[cols="1,1,2,4"]\n|===\n')
    f.write('| Nom | Prénom | Entreprise | Sujet\n')
    for n, fn, c, s, e,w in sorted(zip(names, firstnames, codes, sujets, entreprises,www_entreprises), key=lambda x: x[0].value):
        if c.value == master :
            writeTableEntry(f, n, fn, s, e, w)
    f.write('\n|===')
    f.close()

encadrants = ws['CA']

emails={'m1':' ','m2':' '}
for module, master in masterlist:
    for n,c, e in zip(names,codes,encadrants):
        if c.value == master:
            emails[module]+=e.value.strip()+','

print('M1:',emails['m1'])
print('M2:',emails['m2'])
